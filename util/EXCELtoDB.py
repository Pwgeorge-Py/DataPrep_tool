import pandas as pd
import os
import openpyxl


class MergedData:

    def __init__(self, path, table_name, headers, m_col):
        self.path = path
        self.table_name = table_name
        self.headers = headers
        self.m_col = m_col

    def populate_table(self, db, cursor, module='merged'):
        print("Populating table for " + self.table_name + "...")
        read_file = pd.read_excel(self.path, skiprows=[1, 2, 3, 4, 5, 6], comment="#", names=self.headers, indexcol=0,
                                  usecols=range(self.m_col))
        # Add module column to DataFrame so we know which module the data is from
        read_file['Module'] = module
        read_file.to_sql(self.table_name, db, index=False, if_exists="append")
        print(self.table_name + " added to database")

    def get_col_names(self, cursor):  # Extracting column names from database
        cursor.execute("""SELECT *
                       FROM """ + self.table_name)
        names = list(map(lambda x: x[0], cursor.description))
        with open("Output\col_names.txt", "a") as colFile:
            colFile.write("Column names for table: " + self.table_name + "\n")
            for row in names:
                colFile.write(str(row) + ",")
            colFile.write("\n\n\n")


# --------------------------------------------------------
# --------------------------------------------------------
# --------------------------------------------------------


def get_table_name_and_columns(path):
    wb_obj = openpyxl.load_workbook(path)  # load workboook
    sheet_obj = wb_obj.active  # select sheet

    m_col = sheet_obj.max_column
    m_row = sheet_obj.max_row

    f = os.path.basename(path)  # Get filename
    table_name = f[:f.find(".")]  # Removing filetype from filename to use as table name

    headers = []
    unique_headers = {}

    for r in range(1, m_col + 1):  # Extracting column names
        cell_obj = sheet_obj.cell(row=6, column=r)
        cell_val = cell_obj.value
        if cell_val not in headers:
            headers.append(str(cell_val))
            unique_headers[cell_val] = 1
        else:
            headers.append(str(cell_val + str(unique_headers[cell_val])))
            unique_headers[cell_val] += 1
    return table_name, headers, m_col


def load_all(cursor, db):  # loads all worksheet data into database

    path_list = [r'Input\HS\TimetableModel\HSInterLocationPathSegment_Data.xlsx',
                 r'Input\HS\TimetableModel\HSPlanningPoint_Data.xlsx',
                 r'Input\HS\TrackModel\HSBasePointOfRoute_Data.xlsx',
                 r'Input\HS\TrackModel\HSBasePointsGroup_Data.xlsx',
                 r'Input\HS\TrackModel\HSPlatform_Data.xlsx',
                 r'Input\HS\TrackModel\HSRoute_Data.xlsx',
                 r'Input\HS\TrackModel\HSTrackCircuit_Data.xlsx',
                 r'Input\HS\TimetableCode\HSLineCode_Data.xlsx',
                 r'Input\HS\TimetableCode\HSLocation_Data.xlsx',
                 r'Input\HS\TimetableCode\HSPlatformCode_Data.xlsx',
                 r'Input\HS\TDIF\HSBerth_Data.xlsx',
                 r'Input\HS\AreaIF\UKBerthReplace_Data.xlsx',
                 r'Input\HS\AreaIF\UKBerthStep_Data.xlsx']

    for c, value in enumerate(path_list, 0):
        table_name, headers, m_col = get_table_name_and_columns(path_list[c])
        c = MergedData(path_list[c], table_name, headers, m_col)
        MergedData.populate_table(c, db, cursor)
        c.get_col_names(cursor)


def load_all_unmerged(cursor, db):  # loads all worksheet data into database - from each module

    module_list = ["Ashford", "HL17", "KX", "LB", "Streatham", "TB 2-6", "TB 1abc", "TL", "VC10", "VC 389", "WH"]

    path_start = [r'Input\unmerged\HS&CI - ASHFORD WS1\HS',
                  r'Input\unmerged\HS&CI - HL17\HSData',
                  r'Input\unmerged\HS&CI - KX\HS',
                  r'Input\unmerged\HS&CI - LB\HSData',
                  r'Input\unmerged\HS&CI - ST\HS',
                  r'Input\unmerged\HS&CI - TB - LSE\HS',
                  r'Input\unmerged\HS&CI - TB ASC\HS',
                  r'Input\unmerged\HS&CI - TL\HS',
                  r'Input\unmerged\HS&CI - VC10\HS',
                  r'Input\unmerged\HS&CI - VC\HS',
                  r'Input\unmerged\HS&CI - WH\HS']

    path_end = [r'\TimetableModel\HSInterLocationPathSegment_Data.xlsx',
                r'\TimetableModel\HSPlanningPoint_Data.xlsx',
                r'\TrackModel\HSBasePointOfRoute_Data.xlsx',
                r'\TrackModel\HSBasePointsGroup_Data.xlsx',
                r'\TrackModel\HSPlatform_Data.xlsx',
                r'\TrackModel\HSRoute_Data.xlsx',
                r'\TrackModel\HSTrackCircuit_Data.xlsx',
                r'\TimetableCode\HSLineCode_Data.xlsx',
                r'\TimetableCode\HSLocation_Data.xlsx',
                r'\TimetableCode\HSPlatformCode_Data.xlsx',
                r'\TDIF\HSBerth_Data.xlsx',
                r'\AreaIF\UKBerthReplace_Data.xlsx',
                r'\AreaIF\UKBerthStep_Data.xlsx']

    for c, end_item in enumerate(path_end, 0):
        path = path_start[0] + end_item
        # This extracts the table_name, headers, m_col from the file type then uses it on each file
        table_name, headers, m_col = get_table_name_and_columns(path)
        for z, start_item in enumerate(path_start, 0):
            module = module_list[z]
            print(module)
            path = start_item + end_item
            instance_name = str(c) + str(z)
            instance_name = MergedData(path, table_name, headers, m_col)
            MergedData.populate_table(instance_name, db, cursor, module)
            instance_name.get_col_names(cursor)
